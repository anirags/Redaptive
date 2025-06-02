# ## Required Imports
import os
import warnings
import json
import re
from PyPDF2 import PdfReader, PdfWriter
from dotenv import load_dotenv
from langchain_community.document_loaders import UnstructuredFileLoader
import google.generativeai as genai
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from openpyxl import load_workbook
import pandas as pd
import shutil
import threading
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Ensure correct path resolution
base_dir = os.path.dirname(os.path.abspath(__file__))
warnings.filterwarnings("ignore")
load_dotenv()

# ## Configure Gemini API
genai.configure(api_key=os.getenv("GOOGLE_API_KEY_V2"))
model = genai.GenerativeModel("gemini-1.5-flash")

# ## Folders
input_folder = os.path.join(base_dir, '..', 'temp_in_rwc')
output_folder = os.path.join(base_dir, '..', 'temp_out_rwc')
os.makedirs(output_folder, exist_ok=True)

# Thread lock for Excel writing
excel_lock = threading.Lock()

# ## Extract First Page from PDF
def extract_first_page(input_path, output_path):
    try:
        reader = PdfReader(input_path)
        writer = PdfWriter()
        if reader.pages:
            writer.add_page(reader.pages[0])
            with open(output_path, 'wb') as f_out:
                writer.write(f_out)
        return output_path
    except Exception as e:
        return f"Error processing {input_path}: {str(e)}"

# ## Extract text from a single PDF using OCR
def extract_text_from_pdf(input_path):
    loader = UnstructuredFileLoader(input_path, mode="elements", strategy="ocr_only")
    docs = loader.load()
    return '\n'.join([doc.page_content for doc in docs])

# ## Read Prompt
def get_prompt():
    prompt_path = os.path.join(base_dir, 'rwc_invoice_kpi_prompt.txt')
    with open(prompt_path, 'r') as f:
        return f.read()

# ## Extract KPIs using Gemini
def extract_kpis_with_gemini(prompt, invoice_text):
    full_prompt = f"{prompt}\n\n📄 Here is the invoice text:\n\n{invoice_text}"
    try:
        response = model.generate_content(full_prompt)
        match = re.search(r"\{.*\}", response.text, re.DOTALL)
        return json.loads(match.group()) if match else {"error": "No JSON found in response."}
    except Exception as e:
        return {"error": f"Gemini error: {str(e)}"}

# ## Pipeline for a single PDF file
def process_pdf_file(filename, prompt):
    try:
        input_path = os.path.join(input_folder, filename)
        output_path = os.path.join(output_folder, filename)

        # Step 1: Extract first page
        extract_first_page(input_path, output_path)

        # Step 2: OCR text extraction
        invoice_text = extract_text_from_pdf(output_path)
        if not invoice_text.strip():
            return filename, {"error": "Empty or invalid text extracted."}

        # Step 3: Extract KPIs using Gemini
        result = extract_kpis_with_gemini(prompt, invoice_text)
        return filename, result

    except Exception as e:
        return filename, {"error": str(e)}
    

# Clean numeric columns (remove commas, currency symbols, convert to float)
def clean_numeric(val):
    if isinstance(val, str):
        val = val.replace(",", "").replace("£", "").replace("$", "").strip()
        try:
            return float(val)
        except ValueError:
            return None
    return val


def get_date_cols(billing_dates_str):
    print("Before processing : Billing Dates String:", billing_dates_str)

    converted_dates = [
    f"{date_obj.month}/{date_obj.strftime('%d/%y')}"
    for date_obj in [datetime.strptime(d, '%m/%d/%Y') for d in billing_dates_str]   
]
    print("After Processing --------Converted Billing Dates String:", converted_dates)
    
    # Convert to datetime objects
    billing_dates = sorted([datetime.strptime(date, '%m/%d/%y') for date in converted_dates])
    print("After osrting ----Converted Billing Dates:", billing_dates)
    # Initialize result list
    data = []
    
    result_dict = {}
    # Loop through billing dates to generate 'From' and 'To'
    for i in range(len(billing_dates)):
        to_date = billing_dates[i]
        if i == 0:
            from_date = to_date - relativedelta(months=1)
        else:
            from_date = billing_dates[i - 1]

        # Calculate number of days (inclusive of both ends)
        no_of_days = abs((to_date - from_date).days ) # add 1 if you want to include both start and end dates

        # Format dates manually without leading zeros
        billing_str = f"{to_date.month}/{to_date.day}/{to_date.strftime('%y')}"
        from_str = f"{from_date.month}/{from_date.day}/{from_date.strftime('%y')}"
        to_str = billing_str

        # # Append the row with number of days
        # data.append({
        #     'Billing Date': billing_str,
        #     'From': from_str,
        #     'To': to_str,
        #     'No of Days': no_of_days
        # })

        # Store using the original unsorted full format string (to match input)
        original_key = to_date.strftime('%m/%d/%Y')

        result_dict[original_key] = {
            'From': from_str,
            'To': to_str,
            'No of Days': no_of_days
        }

    # # Convert to DataFrame
    # df = pd.DataFrame(data)
    # return df
    return result_dict

# Alternative approach: Collect all data first, then write once
def write_all_kpis_to_excel(template_path, output_path, all_kpi_data):
    """
    Alternative function to write all KPI data at once instead of individually
    """
    if not all_kpi_data:
        print("No valid KPI data to write")
        return
    
    # Convert all data to DataFrame
    df = pd.DataFrame(all_kpi_data)

    # date_cols = get_date_cols(df['Billing Date'].tolist())

    # date_cols = get_date_cols(df['Billing Date'].tolist())
    date_info_dict = get_date_cols(df['Billing Date'].tolist())

    # print(date_cols)

    # df['Billing Date'] = date_cols['Billing Date']
    # df['From'] = date_cols['From']
    # df['To'] = date_cols['To']
    # df['No of Days'] = date_cols['No of Days']

    # Map each column
    df['From'] = df['Billing Date'].map(lambda x: date_info_dict.get(x, {}).get('From'))
    df['To'] = df['Billing Date'].map(lambda x: date_info_dict.get(x, {}).get('To'))
    df['No of Days'] = df['Billing Date'].map(lambda x: date_info_dict.get(x, {}).get('No of Days'))
    
    
    if df is not None:
        if "City of Cullman Tax $" in df.columns:
            df["City of Cullman Tax $ (2)"] = (df['City of Cullman Tax $']) / (df['Total $ amount'] - df['City of Cullman Tax $']) * 100
            df['City of Cullman Tax $ (2)'] = df['City of Cullman Tax $ (2)'].round(2)
    
        if "Alabama State Taxes $" in df.columns:
            df["Alabama State Taxes $ (2)"] = (df['Alabama State Taxes $']) / (df['Total $ amount'] - df['Alabama State Taxes $']) * 100
            df['Alabama State Taxes $ (2)'] = df['Alabama State Taxes $ (2)'].round(2)

        if "Total kWh" in df.columns:
            df["kWh per day"] = df["Total kWh"] / df["No of Days"]
            df["kWh per day"] = df["kWh per day"].round(2)

        if "Blended rate $/kWh" in df.columns:
            df["Blended rate $/kWh"] = df["Blended rate $/kWh"].round(3)


        if 'Total kWh' in df.columns:
            total_kwh = df['Total kWh'].sum()
            df['kWh %'] = (df['Total kWh'] / total_kwh * 100).round(0).astype(int)
        
        df['Total tax %'] = df['Alabama State Taxes $ (2)'] + df['City of Cullman Tax $ (2)']


    # Source dictionary
    data = {
        "8/30/23": [1500],
        "9/30/23": [1500],
        "10/30/23": [1500],
        "11/30/23": [1500],
        "12/30/23": [1500],
        "1/30/24": [1500],
        "2/29/24": [1500],
        "3/30/24": [1500],
        "4/30/24": [1500, 3888],
        "5/30/24": [1292, 3792],
        "6/30/24": [1532, 4032],
        "7/30/24": [1676, 2736]
    }

    # Convert dictionary to source DataFrame
    df_source = pd.DataFrame.from_dict(data, orient='index')
    df_source.columns = ['Value1', 'Value2']
    df_source = df_source.reset_index().rename(columns={'index': 'Date'})

    # Fill values using condition without merge
    for i, row in df.iterrows():
        date = row['From']
        matched = df_source[df_source['Date'] == date]
        if not matched.empty:
            df.loc[i, 'NC/Max Demand'] = matched.iloc[0]['Value1']
            df.loc[i, 'On-Peak\nDemand'] = matched.iloc[0]['Value2']
            
    column_map ={
        "Billing Date": 2,
        "Month": 3,
        "From":4,
        "To":5,
        "No of Days": 6,
        "Only kWh": 11,
        "Total kWh": 12,
        "kWh per day": 13,
        "kWh %": 14,
        "NC/Max Demand": 15,
        "On-Peak\nDemand": 16,
        "City of Cullman Tax $": 19,
        "Alabama State Taxes $": 20,
        "Alabama State Taxes $ (2)": 21,
        "City of Cullman Tax $ (2)": 22,
        "Total tax %": 23,
        "Current Electric Charges": 25,
        "Total $ amount": 26,
        "Blended rate $/kWh":28
        
        }
    # # Column mapping
    # column_map = {
    #     "Billing Date": 2, "Month": 3, "From": 4, "To": 5, "No of Days": 6,
    #     "Day kWh": 7, "Night kWh": 8, "Total Kwh": 12, "kWh per day": 13,
    #     "KWh": 14, "DUoS Capacity Charge": 15, "Excess Capacity Charge": 16,
    #     "VAT": 19, "Balance Due": 27, "Total $ amount (Without VAT)": 28,
    #     "Blended rate $/kWh (With VAT)": 31, "Blended rate $/kWh (Without VAT)": 32
    # }
    
    # Copy template and write all data
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Write all records starting from row 3
    start_row = 3
    for idx, (_, record) in enumerate(df.iterrows()):
        current_row = start_row + idx
        for kpi, col in column_map.items():
            value = record.get(kpi, "")
            ws.cell(row=current_row, column=col).value = value

    wb.save(output_path)
    print(f"✅ All {len(df)} records written to {output_path}")

    
# ## Run Pipeline - OPTION 2 (Recommended: Collect all data first, then write)
def rwc_run_pipeline_batch_write():
    prompt = get_prompt()
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    template_path = os.path.join(base_dir, '..', 'TemplateExcelOutput_rwc.xlsx')
    output_excel_path = os.path.join(base_dir, '..', 'filled_invoice_rwc.xlsx')

    results = {}
    valid_kpi_data = []
    
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_file = {executor.submit(process_pdf_file, f, prompt): f for f in pdf_files}
        for future in as_completed(future_to_file):
            file = future_to_file[future]
            try:
                fname, res = future.result()
                results[fname] = res
                print(f"📄 Processed ----------->>>>>{fname}: {res}")
                
                # Collect valid KPI data
                if isinstance(res, dict) and not res.get("error"):
                    res['filename'] = fname  # Add filename for reference
                    valid_kpi_data.append(res)
                else:
                    print(f"⚠️ Skipping {fname}: {res.get('error')}")

            except Exception as e:
                results[file] = {"error": f"Unhandled exception: {str(e)}"}

    # Write all valid data to Excel at once
    if valid_kpi_data:
        write_all_kpis_to_excel(template_path, output_excel_path, valid_kpi_data)
    
    return results


# ## Execute
if __name__ == "__main__":
    start_time = time.time()
    
    # Use OPTION 1 (fixed original) or OPTION 2 (batch write - recommended)
    # all_results = run_pipeline()  # Option 1
    all_results = rwc_run_pipeline_batch_write()  # Option 2 - Recommended
    
    end_time = time.time()
    total_time = end_time - start_time
    for file, result in all_results.items():
        print(f"\n🧾 File: {file}\n📊 KPI Result: {result}")

    print(f"\n⏱️ Total processing time: {total_time:.2f} seconds")