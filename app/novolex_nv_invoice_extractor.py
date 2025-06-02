# ## Required Imports
import os
import warnings
import json
import re
from PyPDF2 import PdfReader, PdfWriter
from dotenv import load_dotenv
from langchain_community.document_loaders import UnstructuredFileLoader
import google.generativeai as genai
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from openpyxl import load_workbook
import pandas as pd
import shutil
import threading

# Ensure correct path resolution
base_dir = os.path.dirname(os.path.abspath(__file__))
warnings.filterwarnings("ignore")
load_dotenv()

# ## Configure Gemini API
genai.configure(api_key=os.getenv("GOOGLE_API_KEY_V1"))
model = genai.GenerativeModel("gemini-1.5-flash")

# ## Folders
input_folder = os.path.join(base_dir, '..', 'temp_in_novolex')
output_folder = os.path.join(base_dir, '..', 'temp_out_novolex')
os.makedirs(output_folder, exist_ok=True)

# Thread lock for Excel writing
excel_lock = threading.Lock()

# ## Extract First Page from PDF
def extract_first_page(input_path, output_path):
    # try:
    #     reader = PdfReader(input_path)
    #     writer = PdfWriter()
    #     if reader.pages:
    #         writer.add_page(reader.pages[0])
    #         with open(output_path, 'wb') as f_out:
    #             writer.write(f_out)
    #     return output_path
    # except Exception as e:
    #     return f"Error processing {input_path}: {str(e)}"
    try:
        reader = PdfReader(input_path)
        writer = PdfWriter()
        num_pages_to_extract = min(2, len(reader.pages))  # Avoid IndexError if file has less than 2 pages

        for i in range(num_pages_to_extract):
            writer.add_page(reader.pages[i])

        with open(output_path, 'wb') as f_out:
            writer.write(f_out)

        return output_path
    except Exception as e:
        return f"Error processing {input_path}: {str(e)}"

# ## Extract text from a single PDF using OCR
def extract_text_from_pdf(input_path):
    loader = UnstructuredFileLoader(input_path, mode="elements", strategy="ocr_only")
    docs = loader.load()
    return '\n'.join([doc.page_content for doc in docs])

# ## Read Prompt
def get_prompt():
    prompt_path = os.path.join(base_dir, 'novolex_invoice_kpi_prompt.txt')
    with open(prompt_path, 'r') as f:
        return f.read()

# ## Extract KPIs using Gemini
def extract_kpis_with_gemini(prompt, invoice_text):
    full_prompt = f"{prompt}\n\n📄 Here is the invoice text:\n\n{invoice_text}"
    try:
        response = model.generate_content(full_prompt)
        match = re.search(r"\{.*\}", response.text, re.DOTALL)
        return json.loads(match.group()) if match else {"error": "No JSON found in response."}
    except Exception as e:
        return {"error": f"Gemini error: {str(e)}"}

# ## Pipeline for a single PDF file
def process_pdf_file(filename, prompt):
    try:
        input_path = os.path.join(input_folder, filename)
        output_path = os.path.join(output_folder, filename)

        # Step 1: Extract first page
        extract_first_page(input_path, output_path)

        # Step 2: OCR text extraction
        invoice_text = extract_text_from_pdf(output_path)
        if not invoice_text.strip():
            return filename, {"error": "Empty or invalid text extracted."}

        # Step 3: Extract KPIs using Gemini
        result = extract_kpis_with_gemini(prompt, invoice_text)
        return filename, result

    except Exception as e:
        return filename, {"error": str(e)}
    

# Clean numeric columns (remove commas, currency symbols, convert to float)
def clean_numeric(val):
    if isinstance(val, str):
        val = val.replace(",", "").replace("£", "").replace("$", "").strip()
        try:
            return float(val)
        except ValueError:
            return None
    return val

def write_kpis_to_excel(template_path, output_path, kpi_data):
    # Use thread lock to prevent concurrent access to Excel file
    with excel_lock:
        # 1. Convert dictionary to DataFrame
        df = pd.DataFrame([kpi_data])  # Wrap in list since each call passes single KPI dict

        # Ensure date parsing, handle if keys not present
        if "Billing Date" in df.columns:
                df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce').dt.strftime("%d/%m/%y")
        for col in ["From", "To"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], format="%d/%m/%y", errors='coerce').dt.strftime("%d/%m/%y")
        
        numeric_cols = [
            "Day kWh", "Night kWh", "DUoS Capacity Charge", "Excess Capacity Charge", "VAT", 
            "Total Invoice value"
        ]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_numeric)

        # 2. Derived columns only if date cols exist
        if all(col in df.columns for col in ["Billing Date", "From", "To"]):
            # Convert back to datetime for calculations, then format for display
            billing_dt = pd.to_datetime(df["Billing Date"], format="%d/%m/%y", errors='coerce')
            df["Month"] = billing_dt.dt.strftime("%b-%y")
        
            # For day calculation, convert dates back to datetime with correct format
            from_dt = pd.to_datetime(df['From'], format="%d/%m/%y", errors='coerce')
            to_dt = pd.to_datetime(df['To'], format="%d/%m/%y", errors='coerce')
            df["No of Days"] = (to_dt - from_dt).dt.days + 1
            print("From dates:", df['From'].values)
            print("To dates:", df['To'].values)
            print("From datetime:", from_dt.values)
            print("To datetime:", to_dt.values)
            print("No of days:", df["No of Days"].values)
        else:
            df["Month"] = None
            df["No of Days"] = None

        # kWh per day calculation if columns exist
        if all(col in df.columns for col in ["Day kWh", "Night kWh", "No of Days"]):
            df["kWh per day"] = (df["Day kWh"] + df["Night kWh"]) / df["No of Days"].replace(0, 1)
            # df['kWh per day'] = df['kWh per day'].map(lambda x: f"{x:.2f}")
        else:
            df["kWh per day"] = None

        # Total amount without VAT
        if all(col in df.columns for col in ["Total Invoice value", "VAT"]):
            df["Total $ amount (Without VAT)"] = df["Total Invoice value"] - df["VAT"]
        else:
            df["Total $ amount (Without VAT)"] = None

        total_kwh = None
        if all(col in df.columns for col in ["Day kWh", "Night kWh"]):
            total_kwh = df["Day kWh"] + df["Night kWh"]
        if total_kwh is not None:
            df["Total Kwh"] = total_kwh

        if total_kwh is not None and "Total Invoice value" in df.columns:
            df["Blended rate $/kWh (With VAT)"] = df["Total Invoice value"] / total_kwh.replace(0, 1)
            # df['Blended rate $/kWh (With VAT)'] = df['Blended rate $/kWh (With VAT)'].map(lambda x: f"{x:.2f}")

        else:
            df["Blended rate $/kWh (With VAT)"] = None

        if total_kwh is not None and "Total $ amount (Without VAT)" in df.columns:
            df["Blended rate $/kWh (Without VAT)"] = df["Total $ amount (Without VAT)"] / total_kwh.replace(0, 1)
            # df['Blended rate $/kWh (Without VAT)'] = df['Blended rate $/kWh (Without VAT)'].map(lambda x: f"{x:.2f}")

        else:
            df["Blended rate $/kWh (Without VAT)"] = None

        for col in df.columns:
            print(f"Column: {col}, Type: {df[col].values}")
        
        # Step 3: Column Mapping
        column_map = {
            "Billing Date": 2,
            "Month": 3,
            "From":4,
            "To": 5,
            "No of Days": 6,
            "Day kWh": 7,
            "Night kWh": 8,
            "Total Kwh":12,
            "kWh per day": 13,
            "KWh": 14,
            "DUoS Capacity Charge": 15,
            "Excess Capacity Charge": 16,
            "VAT": 19,
            "Total Invoice value": 27,
            "Total $ amount (Without VAT)": 28,
            "Blended rate $/kWh (With VAT)": 31,
            "Blended rate $/kWh (Without VAT)": 32
        }
        
        # Step 4: Initialize Excel file only once (check if it exists)
        if not os.path.exists(output_path):
            shutil.copy(template_path, output_path)
            print(f"📋 Template copied to {output_path}")
        
        # Load the workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Find next empty row
        row = 3
        while ws.cell(row=row, column=1).value:
            row += 1
            
        # Write data for this record
        for _, record in df.iterrows():
            print(f"Writing row {row} with data: {record.to_dict()}")
            for kpi, col in column_map.items():
                value = record.get(kpi, "")
                ws.cell(row=row, column=col).value = value
            row += 1  # Move to next row for next record

        # Step 6: Save output
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"✅ KPIs written to {output_path} at row {row-1}.")


# Alternative approach: Collect all data first, then write once
def write_all_kpis_to_excel(template_path, output_path, all_kpi_data):
    """
    Alternative function to write all KPI data at once instead of individually
    """
    if not all_kpi_data:
        print("No valid KPI data to write")
        return
    
    # Convert all data to DataFrame
    df = pd.DataFrame(all_kpi_data)
    
    # Process all the data transformations (same as before)
    if "Billing Date" in df.columns:
        # df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce').dt.strftime("%m/%d/%y")
        df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce')

    for col in ["From", "To"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format="%d/%m/%y", errors='coerce').dt.strftime("%d/%m/%y")
    
    numeric_cols = [
        "Day kWh", "Night kWh", "DUoS Capacity Charge", "Excess Capacity Charge", "VAT", 
        "Total Invoice value"
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)

    # Derived columns
    if all(col in df.columns for col in ["Billing Date", "From", "To"]):
        # Convert back to datetime for calculations, then format for display
        # billing_dt = pd.to_datetime(df["Billing Date"], format="%d/%m/%y", errors='coerce')
        df["Month"] = df["Billing Date"].dt.strftime("%b-%y")
        df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce').dt.strftime("%m/%d/%y")

        
        # For day calculation, convert dates back to datetime with correct format
        from_dt = pd.to_datetime(df['From'], format="%d/%m/%y", errors='coerce')
        to_dt = pd.to_datetime(df['To'], format="%d/%m/%y", errors='coerce')
        df["No of Days"] = (to_dt - from_dt).dt.days
    else:
        df["Month"] = None
        df["No of Days"] = None

    if all(col in df.columns for col in ["Day kWh", "Night kWh", "No of Days"]):
        df["kWh per day"] = (df["Day kWh"] + df["Night kWh"]) / df["No of Days"].replace(0, 1)
        df["kWh per day"] = df["kWh per day"].round(1)
        
    else:
        df["kWh per day"] = None

    if all(col in df.columns for col in ["Total Invoice value", "VAT"]):
        df["Total $ amount (Without VAT)"] = df["Total Invoice value"] - df["VAT"]
       
    else:
        df["Total $ amount (Without VAT)"] = None

    total_kwh = None
    if all(col in df.columns for col in ["Day kWh", "Night kWh"]):
        total_kwh = df["Day kWh"] + df["Night kWh"]
        df["Total Kwh"] = total_kwh
        

    if total_kwh is not None and "Total Invoice value" in df.columns:
        df["Blended rate $/kWh (With VAT)"] = df["Total Invoice value"] / total_kwh.replace(0, 1)
        df["Blended rate $/kWh (With VAT)"] = df["Blended rate $/kWh (With VAT)"].round(2)
        
    else:
        df["Blended rate $/kWh (With VAT)"] = None

    if total_kwh is not None and "Total $ amount (Without VAT)" in df.columns:
        df["Blended rate $/kWh (Without VAT)"] = df["Total $ amount (Without VAT)"] / total_kwh.replace(0, 1)
        df["Blended rate $/kWh (Without VAT)"] = df["Blended rate $/kWh (Without VAT)"].round(2)
        
    else:
        df["Blended rate $/kWh (Without VAT)"] = None

    if 'Total Kwh' in df.columns:
            total_kwh = df['Total Kwh'].sum()
            df['kWh %'] = (df['Total Kwh'] / total_kwh * 100)
            df['kWh %'] = df['kWh %'].round(0)
    # if 'Total Kwh' in df.columns:
    #         total_kwh = df['Total Kwh'].sum()
    #         df['kWh %'] = df['Total Kwh'] / total_kwh
    if "VAT" in df.columns:
        df['VAT'] = df['VAT'].round(2)


    

    # Column mapping
    column_map = {
        "Billing Date": 2, "Month": 3, "From": 4, "To": 5, "No of Days": 6,
        "Day kWh": 7, "Night kWh": 8, "Total Kwh": 12, "kWh per day": 13,
        "kWh %": 14, "DUoS Capacity Charge": 15, "Excess Capacity Charge": 16,
        "VAT": 19, "Total Invoice value": 27, "Total $ amount (Without VAT)": 28,
        "Blended rate $/kWh (With VAT)": 31, "Blended rate $/kWh (Without VAT)": 32
    }
    
    # Copy template and write all data
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Write all records starting from row 3
    start_row = 3
    for idx, (_, record) in enumerate(df.iterrows()):
        current_row = start_row + idx
        for kpi, col in column_map.items():
            value = record.get(kpi, "")
            ws.cell(row=current_row, column=col).value = value

    wb.save(output_path)
    print(f"✅ All {len(df)} records written to {output_path}")

    


# ## Run Pipeline on All PDFs in Parallel - OPTION 1 (Fixed original approach)
def run_pipeline():
    prompt = get_prompt()
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    template_path = os.path.join(base_dir, '..', 'TemplateExcelOutput_novolex.xlsx')
    output_excel_path = os.path.join(base_dir, '..', 'filled_invoice_novolex.xlsx')

    results = {}
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_file = {executor.submit(process_pdf_file, f, prompt): f for f in pdf_files}
        for future in as_completed(future_to_file):
            file = future_to_file[future]
            try:
                fname, res = future.result()
                results[fname] = res
                print(f"📄 Processed ----------->>>>>{fname}: {res}")
                # ✅ Write to Excel only if no error
                if isinstance(res, dict) and not res.get("error"):
                    write_kpis_to_excel(template_path, output_excel_path, res)
                else:
                    print(f"⚠️ Skipping Excel write for {fname}: {res.get('error')}")

            except Exception as e:
                results[file] = {"error": f"Unhandled exception: {str(e)}"}

    return results


# ## Run Pipeline - OPTION 2 (Recommended: Collect all data first, then write)
def novolex_run_pipeline_batch_write():
    prompt = get_prompt()
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    template_path = os.path.join(base_dir, '..', 'TemplateExcelOutput_novolex.xlsx')
    output_excel_path = os.path.join(base_dir, '..', 'filled_invoice_novolex.xlsx')

    results = {}
    valid_kpi_data = []
    
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_file = {executor.submit(process_pdf_file, f, prompt): f for f in pdf_files}
        for future in as_completed(future_to_file):
            file = future_to_file[future]
            try:
                fname, res = future.result()
                results[fname] = res
                print(f"📄 Processed ----------->>>>>{fname}: {res}")
                
                # Collect valid KPI data
                if isinstance(res, dict) and not res.get("error"):
                    res['filename'] = fname  # Add filename for reference
                    valid_kpi_data.append(res)
                else:
                    print(f"⚠️ Skipping {fname}: {res.get('error')}")

            except Exception as e:
                results[file] = {"error": f"Unhandled exception: {str(e)}"}

    # Write all valid data to Excel at once
    if valid_kpi_data:
        write_all_kpis_to_excel(template_path, output_excel_path, valid_kpi_data)
    
    return results


# ## Execute
if __name__ == "__main__":
    start_time = time.time()
    
    # Use OPTION 1 (fixed original) or OPTION 2 (batch write - recommended)
    # all_results = run_pipeline()  # Option 1
    all_results = novolex_run_pipeline_batch_write()  # Option 2 - Recommended
    
    end_time = time.time()
    total_time = end_time - start_time
    for file, result in all_results.items():
        print(f"\n🧾 File: {file}\n📊 KPI Result: {result}")

    print(f"\n⏱️ Total processing time: {total_time:.2f} seconds")